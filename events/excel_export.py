"""
Gestion du fichier Excel des inscrits, un fichier par événement.

Le fichier est régénéré à chaque inscription confirmée, à partir de la
base de données (source de vérité). Cela garantit que le fichier Excel
reste toujours synchronisé avec les inscriptions confirmées, même en
cas d'erreur ponctuelle d'écriture lors d'une inscription précédente.
"""
import re
from pathlib import Path

# Django et openpyxl sont des dépendances optionnelles, utilisées uniquement pour la génération du fichier Excel. Nous les importons localement dans les fonctions
# qui en ont besoin, afin d'éviter de les charger inutilement dans d'autres contextes (ex: lors de l'exécution de tests qui ne touchent pas à l'Excel).
from django.conf import settings 
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


COLONNES = ["Nom", "Prénoms", "Téléphone", "Ville", "Email", "Date de confirmation"]

# Le dossier "inscrits" est créé dans MEDIA_ROOT s'il n'existe pas déjà, et les fichiers Excel y sont stockés. Cela permet de garder les fichiers générés à l'écart du code source et de les servir facilement via le serveur de fichiers statiques de Django si besoin.
def _dossier_inscrits(): 
    dossier = Path(settings.MEDIA_ROOT) / "inscrits"
    dossier.mkdir(parents=True, exist_ok=True)
    return dossier


def _nom_fichier_securise(titre_evenement, evenement_id):
    """Construit un nom de fichier sûr à partir du titre de l'événement."""
    slug = re.sub(r"[^\w\s-]", "", titre_evenement).strip().lower()
    slug = re.sub(r"[\s]+", "_", slug)[:60]
    return f"inscrits_{evenement_id}_{slug}.xlsx"


def chemin_fichier_excel(evenement):
    return _dossier_inscrits() / _nom_fichier_securise(evenement.titre, evenement.pk)


def regenerer_excel_evenement(evenement):
    """Régénère entièrement le fichier Excel des inscrits confirmés
    pour un événement donné, à partir de la base de données.

    Retourne le chemin du fichier créé/mis à jour.
    """
    from .models import Inscription  # import local pour éviter une dépendance circulaire

    inscrits = (
        evenement.inscriptions
        .filter(statut=Inscription.Statut.CONFIRMEE)
        .order_by("date_confirmation")
    )

    classeur = Workbook()
    feuille = classeur.active
    feuille.title = "Inscrits"

    # En-tête
    feuille.append([evenement.titre])
    feuille.append([f"{evenement.lieu} — du {evenement.date_debut:%d/%m/%Y} au {evenement.date_fin:%d/%m/%Y}"])
    feuille.append([])
    feuille.append(COLONNES)

    titre_cell = feuille.cell(row=1, column=1)
    titre_cell.font = Font(bold=True, size=14)

    ligne_entetes = feuille[4]
    for cell in ligne_entetes:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2E5C8A", end_color="2E5C8A", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    for inscription in inscrits:
        feuille.append([
            inscription.nom,
            inscription.prenoms,
            inscription.telephone1,
            inscription.ville,
            inscription.email,
            inscription.date_confirmation.strftime("%d/%m/%Y %H:%M") if inscription.date_confirmation else "",
        ])

    # Largeur des colonnes
    largeurs = [18, 22, 16, 18, 30, 20]
    for i, largeur in enumerate(largeurs, start=1):
        feuille.column_dimensions[get_column_letter(i)].width = largeur

    chemin = chemin_fichier_excel(evenement)
    classeur.save(chemin)
    return chemin
